"""
Excel文件拆分工具
支持按指定列拆分Excel文件，保留所有sheet结构
"""
import pandas as pd
import os
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelSplitter:
    """Excel文件拆分器"""
    
    def __init__(self, input_file: str, split_column: str, output_dir: str = "output"):
        """
        初始化拆分器
        
        Args:
            input_file: 输入的Excel文件路径
            split_column: 用于拆分的列名（如"商务组别"）
            output_dir: 输出目录路径
        """
        self.input_file = input_file
        self.split_column = split_column
        self.output_dir = output_dir
        
        # 创建输出目录
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
    def read_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """
        读取Excel文件中的所有sheet
        
        Returns:
            字典，key为sheet名称，value为DataFrame
        """
        excel_file = pd.ExcelFile(self.input_file)
        sheets = {}
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(self.input_file, sheet_name=sheet_name)
            sheets[sheet_name] = df
            
        return sheets
    
    def get_unique_values(self, sheets: Dict[str, pd.DataFrame]) -> List[str]:
        """
        获取所有sheet中指定列的唯一值
        
        Args:
            sheets: 所有sheet的字典
            
        Returns:
            唯一值列表
        """
        all_values = set()
        
        for sheet_name, df in sheets.items():
            if self.split_column in df.columns:
                # 去除空值并添加到集合
                values = df[self.split_column].dropna().unique()
                all_values.update(values)
            else:
                print(f"警告: Sheet '{sheet_name}' 中未找到列 '{self.split_column}'")
        
        return sorted(list(all_values))
    
    def copy_sheet_formatting(self, source_wb, source_sheet_name: str, 
                            target_wb, target_sheet_name: str):
        """
        复制sheet的格式（尽可能保留原始格式）
        
        Args:
            source_wb: 源工作簿
            source_sheet_name: 源sheet名称
            target_wb: 目标工作簿
            target_sheet_name: 目标sheet名称
        """
        try:
            source_ws = source_wb[source_sheet_name]
            target_ws = target_wb[target_sheet_name]
            
            # 复制列宽
            for col in source_ws.column_dimensions:
                if col in source_ws.column_dimensions:
                    target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width
            
            # 复制行高
            for row in source_ws.row_dimensions:
                if row in source_ws.row_dimensions:
                    target_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height
                    
        except Exception as e:
            print(f"复制格式时出错: {str(e)}")
    
    def split_and_save(self) -> Dict[str, str]:
        """
        执行拆分并保存文件
        
        Returns:
            字典，key为拆分值，value为生成的文件路径
        """
        # 读取所有sheet
        print(f"正在读取文件: {self.input_file}")
        sheets = self.read_all_sheets()
        print(f"共找到 {len(sheets)} 个sheet")
        
        # 获取唯一值
        unique_values = self.get_unique_values(sheets)
        print(f"在列 '{self.split_column}' 中找到 {len(unique_values)} 个唯一值")
        
        if not unique_values:
            raise ValueError(f"未找到可用于拆分的数据。请检查列名 '{self.split_column}' 是否正确。")
        
        # 加载原始工作簿用于复制格式
        try:
            source_wb = openpyxl.load_workbook(self.input_file)
        except:
            source_wb = None
        
        # 为每个唯一值创建新的Excel文件
        output_files = {}
        
        for value in unique_values:
            # 清理文件名中的非法字符
            safe_filename = str(value).replace('/', '_').replace('\\', '_').replace(':', '_')
            output_file = os.path.join(self.output_dir, f"{safe_filename}.xlsx")
            
            print(f"正在创建文件: {safe_filename}.xlsx")
            
            # 创建新的Excel写入器
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                sheet_written = False
                
                # 遍历所有sheet
                for sheet_name, df in sheets.items():
                    if self.split_column in df.columns:
                        # 筛选当前值的数据
                        filtered_df = df[df[self.split_column] == value]
                        
                        if not filtered_df.empty:
                            # 写入数据
                            filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            sheet_written = True
                            print(f"  - Sheet '{sheet_name}': {len(filtered_df)} 行数据")
                
                if sheet_written:
                    # 尝试复制格式
                    if source_wb:
                        for sheet_name in sheets.keys():
                            if sheet_name in writer.book.sheetnames:
                                try:
                                    self.copy_sheet_formatting(
                                        source_wb, sheet_name,
                                        writer.book, sheet_name
                                    )
                                except:
                                    pass
            
            if sheet_written:
                output_files[value] = output_file
                print(f"✓ 成功创建: {safe_filename}.xlsx")
            else:
                # 如果没有写入任何sheet，删除空文件
                if os.path.exists(output_file):
                    os.remove(output_file)
        
        return output_files
    
    def get_summary(self) -> str:
        """
        获取拆分摘要信息
        
        Returns:
            摘要字符串
        """
        sheets = self.read_all_sheets()
        unique_values = self.get_unique_values(sheets)
        
        summary = f"""
拆分配置摘要:
-----------------
输入文件: {self.input_file}
拆分列: {self.split_column}
Sheet数量: {len(sheets)}
Sheet名称: {', '.join(sheets.keys())}
将生成文件数: {len(unique_values)}
拆分值: {', '.join(str(v) for v in unique_values)}
输出目录: {self.output_dir}
"""
        return summary


def main():
    """命令行使用示例"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel文件拆分工具')
    parser.add_argument('input_file', help='输入的Excel文件路径')
    parser.add_argument('split_column', help='用于拆分的列名（如"商务组别"）')
    parser.add_argument('--output-dir', '-o', default='output', help='输出目录（默认: output）')
    
    args = parser.parse_args()
    
    # 创建拆分器
    splitter = ExcelSplitter(
        input_file=args.input_file,
        split_column=args.split_column,
        output_dir=args.output_dir
    )
    
    # 显示摘要
    print(splitter.get_summary())
    
    # 执行拆分
    print("\n开始拆分...")
    output_files = splitter.split_and_save()
    
    print(f"\n拆分完成！共生成 {len(output_files)} 个文件。")
    print(f"文件保存在: {os.path.abspath(args.output_dir)}")


if __name__ == '__main__':
    main()
